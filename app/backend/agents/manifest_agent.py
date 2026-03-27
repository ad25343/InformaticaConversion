# Copyright (c) 2026 ad25343 — https://github.com/ad25343/InformaticaConversion
# Licensed under CC BY-NC 4.0. Commercial use requires written permission.
"""
STEP 1.5 — Mapping Manifest Agent
Runs immediately after XML parsing, before conversion.

Analyses the graph dict produced by parser_agent and builds a structured
ManifestReport that surfaces every source-to-target connection with a
confidence score, plus all gaps and ambiguous items the reviewer must resolve.

Outputs:
  1. ManifestReport (in-memory, returned to caller)
  2. manifest_<mapping_name>.xlsx  (written to the path supplied by caller)

The xlsx has three sheets:
  • Summary        — one row per mapping: counts and overall status
  • Full Lineage   — every source connection (green=HIGH, amber=MEDIUM, yellow=LOW/UNMAPPED)
  • Review Required — filtered view of LOW + UNMAPPED rows with an editable Override column

When the reviewer fills in the Override column and the xlsx is re-uploaded,
conversion_agent reads it back via load_overrides() and uses those values as
ground truth for any ambiguous or missing connections.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

from ..models.schemas import (
    ManifestConfidence, ManifestItem, ManifestItemType,
    ManifestReport, ManifestOverride,
)


# ─────────────────────────────────────────────────────────────────────────────
# Colour palette
# ─────────────────────────────────────────────────────────────────────────────
_GREEN   = "C6EFCE"   # HIGH confidence
_AMBER   = "FFEB9C"   # MEDIUM confidence
_YELLOW  = "FFC7CE"   # LOW confidence
_RED_TXT = "9C0006"   # LOW text
_BLUE_H  = "1F4E79"   # Header fill (dark blue)
_WHITE   = "FFFFFF"

# Item types included on the Full Lineage sheet (overview)
_LINEAGE_ITEM_TYPES = (
    ManifestItemType.SOURCE_LINEAGE,
    ManifestItemType.LOOKUP,
    ManifestItemType.PARAMETER,
    ManifestItemType.EXPRESSION,
)

# Confidences requiring reviewer action
_REVIEW_CONFIDENCES = (ManifestConfidence.LOW, ManifestConfidence.UNMAPPED)


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def build_manifest(graph: dict) -> ManifestReport:
    """Analyse the parser graph dict and return a ManifestReport."""
    mapping_names, source_names, target_names, all_transformations, all_connectors = (
        _extract_graph_collections(graph)
    )
    sq_context = _build_sq_context(all_transformations, all_connectors)
    connected_instances, sq_connected, lookup_source_names = sq_context

    items: list[ManifestItem] = []
    _collect_source_lineage(
        graph, all_connectors, connected_instances, sq_connected,
        lookup_source_names, mapping_names, items,
    )
    _collect_orphaned_ports(all_transformations, all_connectors, graph, items)
    _collect_lineage_gaps(all_transformations, all_connectors, graph, items)
    _collect_expression_items(all_transformations, graph, items)
    _collect_lookup_items(all_transformations, graph, items)
    _collect_parameter_items(graph, mapping_names, items)

    conf_counts = _count_confidences(items)
    return ManifestReport(
        mapping_names=mapping_names,
        source_count=len(source_names),
        target_count=len(target_names),
        transformation_count=len(all_transformations),
        high_confidence=conf_counts[ManifestConfidence.HIGH],
        medium_confidence=conf_counts[ManifestConfidence.MEDIUM],
        low_confidence=conf_counts[ManifestConfidence.LOW],
        unmapped_count=conf_counts[ManifestConfidence.UNMAPPED],
        review_required=(
            conf_counts[ManifestConfidence.LOW] > 0
            or conf_counts[ManifestConfidence.UNMAPPED] > 0
        ),
        items=items,
        generated_at=datetime.now(timezone.utc).isoformat(),
    )


# ─────────────────────────────────────────────────────────────────────────────
# Graph extraction helpers
# ─────────────────────────────────────────────────────────────────────────────

def _flatten_mapping_key(graph: dict, key: str) -> list:
    """Flatten a per-mapping list key across all mappings."""
    result = []
    for m in graph.get("mappings", []):
        result.extend(m.get(key, []))
    return result


def _extract_graph_collections(graph: dict) -> tuple:
    """Extract the five flat collections from a parser graph dict."""
    mapping_names       = [m["name"] for m in graph.get("mappings", [])]
    source_names        = [s["name"] for s in graph.get("sources", [])]
    target_names        = [t["name"] for t in graph.get("targets", [])]
    all_transformations = _flatten_mapping_key(graph, "transformations")
    all_connectors      = _flatten_mapping_key(graph, "connectors")
    return mapping_names, source_names, target_names, all_transformations, all_connectors


def _sq_names_set(all_transformations: list) -> set:
    """Return the set of Source Qualifier transformation names."""
    return {t["name"] for t in all_transformations if t.get("type") == "Source Qualifier"}


def _lookup_source_map(all_transformations: list) -> dict[str, str]:
    """Build a map of lookup-table-name → LKP transform name."""
    result: dict[str, str] = {}
    for t in all_transformations:
        if t.get("type") != "Lookup":
            continue
        lkp_src = t.get("table_attribs", {}).get("Lookup table name", "").strip()
        if lkp_src:
            result[lkp_src] = t["name"]
    return result


def _build_sq_context(
    all_transformations: list, all_connectors: list
) -> tuple[set, set, dict]:
    """Build the sets and dict needed to score source confidence."""
    connected_instances = (
        {c.get("from_instance") for c in all_connectors}
        | {c.get("to_instance") for c in all_connectors}
    )
    sq_names     = _sq_names_set(all_transformations)
    sq_connected = {n for n in sq_names if n in connected_instances}
    lookup_source_names = _lookup_source_map(all_transformations)
    return connected_instances, sq_connected, lookup_source_names


def _count_confidences(items: list[ManifestItem]) -> dict:
    """Return a confidence → count dict for all items."""
    counts = {c: 0 for c in ManifestConfidence}
    for item in items:
        counts[item.confidence] += 1
    return counts


# ─────────────────────────────────────────────────────────────────────────────
# Item collection helpers
# ─────────────────────────────────────────────────────────────────────────────

def _collect_source_lineage(
    graph: dict, all_connectors: list, connected_instances: set,
    sq_connected: set, lookup_source_names: dict, mapping_names: list,
    items: list,
) -> None:
    for src in graph.get("sources", []):
        src_name = src["name"]
        confidence, determination = _score_source(
            src_name, connected_instances, sq_connected, lookup_source_names
        )
        items.append(ManifestItem(
            mapping_name=_mapping_for_source(src_name, all_connectors, mapping_names),
            item_type=ManifestItemType.SOURCE_LINEAGE,
            location=src_name,
            description=f"Source '{src_name}' connection to downstream qualifier / lookup",
            tool_determination=determination,
            confidence=confidence,
        ))


def _is_skippable_output_port(ttype: str, pname: str, porttype: str) -> bool:
    """Return True for ports that should not be checked for orphan status."""
    if "OUTPUT" not in porttype:
        return True
    if ttype == "Target":
        return True
    if ttype == "Rank" and pname == "RANKINDEX":
        return True
    return False


def _port_from_set(all_connectors: list) -> set:
    """Build the set of (from_instance, from_field) pairs from connectors."""
    return {(c.get("from_instance"), c.get("from_field")) for c in all_connectors}


def _port_to_set(all_connectors: list) -> set:
    """Build the set of (to_instance, to_field) pairs from connectors."""
    return {(c.get("to_instance"), c.get("to_field")) for c in all_connectors}


def _collect_orphaned_ports(
    all_transformations: list, all_connectors: list, graph: dict, items: list
) -> None:
    port_from = _port_from_set(all_connectors)
    for t in all_transformations:
        ttype = t.get("type", "")
        tname = t.get("name", "")
        mapping_name = _mapping_for_transform(tname, graph)
        for port in t.get("ports", []):
            pname    = port.get("name", "")
            porttype = port.get("porttype", "")
            if _is_skippable_output_port(ttype, pname, porttype):
                continue
            if (tname, pname) not in port_from:
                items.append(ManifestItem(
                    mapping_name=mapping_name,
                    item_type=ManifestItemType.ORPHANED_PORT,
                    location=f"{tname}.{pname}",
                    description=(
                        f"Output port '{pname}' on {ttype} '{tname}' has no downstream connector"
                    ),
                    tool_determination="No connector found originating from this port",
                    confidence=ManifestConfidence.LOW,
                    notes="Enter IGNORE if intentionally unmapped, or the correct target port name",
                ))


def _collect_lineage_gaps(
    all_transformations: list, all_connectors: list, graph: dict, items: list
) -> None:
    port_to = _port_to_set(all_connectors)
    for t in all_transformations:
        if t.get("type") != "Target":
            continue
        tname        = t["name"]
        mapping_name = _mapping_for_transform(tname, graph)
        for port in t.get("ports", []):
            pname = port.get("name", "")
            if (tname, pname) not in port_to:
                items.append(ManifestItem(
                    mapping_name=mapping_name,
                    item_type=ManifestItemType.LINEAGE_GAP,
                    location=f"{tname}.{pname}",
                    description=f"Target field '{pname}' on '{tname}' has no incoming connector",
                    tool_determination="No connector found feeding this target field",
                    confidence=ManifestConfidence.LOW,
                    notes="Enter source port (e.g. SQ_X.FIELD_Y) or INTENTIONAL_NULL",
                ))


def _expression_description(expr: str) -> str:
    """Truncate an expression string for display in the manifest."""
    return f"Expression to convert: {expr[:120]}{'…' if len(expr) > 120 else ''}"


def _is_nontrivial_expression(expr: str, port_name: str) -> bool:
    """Return True if expr contains a non-trivial expression worth recording."""
    trivial_values = ("", port_name)
    return bool(expr) and expr.strip() not in trivial_values


def _collect_expression_items(all_transformations: list, graph: dict, items: list) -> None:
    for t in all_transformations:
        if t.get("type") != "Expression":
            continue
        tname        = t["name"]
        mapping_name = _mapping_for_transform(tname, graph)
        for port in t.get("ports", []):
            expr = port.get("expression", "")
            if not _is_nontrivial_expression(expr, port.get("name", "")):
                continue
            items.append(ManifestItem(
                mapping_name=mapping_name,
                item_type=ManifestItemType.EXPRESSION,
                location=f"{tname}.{port['name']}",
                description=_expression_description(expr),
                tool_determination="Will be converted by Claude during conversion step",
                confidence=ManifestConfidence.HIGH,
            ))


def _collect_lookup_items(all_transformations: list, graph: dict, items: list) -> None:
    for t in all_transformations:
        if t.get("type") != "Lookup":
            continue
        tname    = t["name"]
        lkp_src  = t.get("table_attribs", {}).get("Lookup table name", "")
        lkp_cond = t.get("table_attribs", {}).get("Lookup condition", "")
        cond_str = f" — condition: {lkp_cond[:80]}" if lkp_cond else ""
        items.append(ManifestItem(
            mapping_name=_mapping_for_transform(tname, graph),
            item_type=ManifestItemType.LOOKUP,
            location=tname,
            description=f"Lookup against '{lkp_src}'{cond_str}",
            tool_determination=f"Reference table: {lkp_src}",
            confidence=ManifestConfidence.HIGH,
        ))


def _param_confidence(pval: str) -> ManifestConfidence:
    """Return MEDIUM if a default value is set, LOW otherwise."""
    return ManifestConfidence.MEDIUM if pval else ManifestConfidence.LOW


def _collect_parameter_items(graph: dict, mapping_names: list, items: list) -> None:
    default_mapping = mapping_names[0] if mapping_names else "unknown"
    for param in graph.get("parameters", []):
        pname = param.get("name", "")
        pval  = param.get("default_value", "")
        items.append(ManifestItem(
            mapping_name=default_mapping,
            item_type=ManifestItemType.PARAMETER,
            location=pname,
            description=f"Parameter '{pname}' — default: {pval or '(none)'}",
            tool_determination=f"Default value: {pval or 'not set'}",
            confidence=_param_confidence(pval),
            notes="Override with environment-specific value if default is wrong",
        ))


# ─────────────────────────────────────────────────────────────────────────────
# xlsx read-back — load_overrides
# ─────────────────────────────────────────────────────────────────────────────

def write_xlsx(report: ManifestReport, path: str) -> None:
    """Write the manifest report to an xlsx file at *path*."""
    wb = Workbook()
    _build_summary_sheet(wb, report)
    _build_lineage_sheet(wb, report)
    _build_review_sheet(wb, report)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)


def write_xlsx_bytes(report: ManifestReport) -> bytes:
    """Return the xlsx as raw bytes (for in-memory use / API response)."""
    buf = io.BytesIO()
    wb = Workbook()
    _build_summary_sheet(wb, report)
    _build_lineage_sheet(wb, report)
    _build_review_sheet(wb, report)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(buf)
    return buf.getvalue()


def _open_review_sheet(path: str):
    """Open the 'Review Required' worksheet or return None on failure."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
    except Exception:
        return None
    if "Review Required" not in wb.sheetnames:
        return None
    return wb["Review Required"]


def _collect_header_row(row, headers: dict[str, int]) -> None:
    """Add string cell values from one header row into the headers dict."""
    for cell in row:
        if cell.value and isinstance(cell.value, str):
            headers[cell.value.strip()] = cell.column


def _find_override_headers(ws) -> dict[str, int]:
    """Scan the first 5 rows to build a header-name → column-index map."""
    headers: dict[str, int] = {}
    for row in ws.iter_rows(min_row=1, max_row=5):
        _collect_header_row(row, headers)
        if headers:
            break
    return headers


def _parse_item_type(raw) -> ManifestItemType:
    """Convert a raw cell value to a ManifestItemType, defaulting to SOURCE_LINEAGE."""
    try:
        return ManifestItemType(raw) if raw else ManifestItemType.SOURCE_LINEAGE
    except ValueError:
        return ManifestItemType.SOURCE_LINEAGE


def _extract_row_cells(
    row: tuple,
    loc_col: int,
    type_col: Optional[int],
    override_col: int,
    notes_col: Optional[int],
) -> Optional[tuple]:
    """Extract the four cell values from a row tuple, returning None on IndexError."""
    try:
        return (
            row[loc_col - 1],
            row[type_col - 1] if type_col else None,
            row[override_col - 1],
            row[notes_col - 1] if notes_col else None,
        )
    except IndexError:
        return None


def _notes_str(notes) -> Optional[str]:
    """Return stripped notes string or None."""
    return str(notes).strip() if notes else None


def _parse_override_row(
    row: tuple,
    loc_col: int,
    type_col: Optional[int],
    override_col: int,
    notes_col: Optional[int],
) -> Optional[ManifestOverride]:
    """Parse one data row into a ManifestOverride, or return None to skip."""
    cells = _extract_row_cells(row, loc_col, type_col, override_col, notes_col)
    if cells is None:
        return None
    location, item_type_raw, override, notes = cells
    if not location or not override:
        return None
    return ManifestOverride(
        location=str(location).strip(),
        item_type=_parse_item_type(item_type_raw),
        reviewer_override=str(override).strip(),
        notes=_notes_str(notes),
    )


def _get_required_cols(
    headers: dict[str, int],
) -> Optional[tuple[int, int, Optional[int], Optional[int]]]:
    """Return (loc_col, override_col, type_col, notes_col) or None if required columns missing."""
    loc_col      = headers.get("Location")
    override_col = headers.get("Reviewer Override")
    if not loc_col or not override_col:
        return None
    return loc_col, override_col, headers.get("Item Type"), headers.get("Notes")


def _parse_override_rows(ws, cols: tuple) -> list[ManifestOverride]:
    """Iterate worksheet data rows and collect non-None ManifestOverride objects."""
    loc_col, override_col, type_col, notes_col = cols
    overrides: list[ManifestOverride] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = _parse_override_row(row, loc_col, type_col, override_col, notes_col)
        if item is not None:
            overrides.append(item)
    return overrides


def load_overrides(path: str) -> list[ManifestOverride]:
    """
    Read back a reviewer-annotated manifest xlsx and return only the rows
    where the reviewer filled in an Override value.  The conversion agent
    calls this before it begins to resolve ambiguous connections.
    """
    ws = _open_review_sheet(path)
    if ws is None:
        return []

    cols = _get_required_cols(_find_override_headers(ws))
    if cols is None:
        return []

    return _parse_override_rows(ws, cols)


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers — confidence scoring
# ─────────────────────────────────────────────────────────────────────────────

def _score_direct(
    src_name: str, connected_instances: set[str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(a) Direct connector from this source instance."""
    if src_name in connected_instances:
        return ManifestConfidence.HIGH, f"Direct connector from '{src_name}'"
    return None


def _score_exact_sq(
    src_name: str, sq_connected: set[str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(b) Exact SQ match — SQ_SOURCENAME."""
    exact_sq = f"SQ_{src_name}"
    if exact_sq in sq_connected:
        return ManifestConfidence.HIGH, f"Exact SQ match → {exact_sq}"
    return None


def _score_sq_contains_src(
    src_name: str, sq_connected: set[str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(c) Source name contained in some SQ name."""
    for sq in sq_connected:
        if src_name in sq:
            return ManifestConfidence.HIGH, f"Source name found in SQ name → {sq}"
    return None


def _score_src_contains_sq_stem(
    src_name: str, sq_connected: set[str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(d) SQ abbreviated name contained in source name."""
    for sq in sq_connected:
        stem = sq.replace("SQ_", "")
        if stem and stem in src_name:
            return (
                ManifestConfidence.MEDIUM,
                f"SQ stem '{stem}' found in source name (abbreviated match) → {sq}",
            )
    return None


def _score_lookup(
    src_name: str, lookup_source_names: dict[str, str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(e) Lookup reference table."""
    if src_name in lookup_source_names:
        lkp = lookup_source_names[src_name]
        return ManifestConfidence.HIGH, f"Lookup reference table → {lkp}"
    return None


def _score_token_overlap(
    src_name: str, sq_connected: set[str]
) -> Optional[tuple[ManifestConfidence, str]]:
    """(f) Partial token overlap — weak match."""
    src_tokens = set(src_name.upper().split("_"))
    for sq in sq_connected:
        sq_tokens = set(sq.replace("SQ_", "").upper().split("_"))
        if src_tokens & sq_tokens:
            return ManifestConfidence.LOW, f"Partial token overlap with SQ '{sq}' — needs review"
    return None


def _score_source(
    src_name: str,
    connected_instances: set[str],
    sq_connected: set[str],
    lookup_source_names: dict[str, str],
) -> tuple[ManifestConfidence, str]:
    """Return (confidence, human-readable determination string) for a source."""
    scorers = [
        lambda: _score_direct(src_name, connected_instances),
        lambda: _score_exact_sq(src_name, sq_connected),
        lambda: _score_sq_contains_src(src_name, sq_connected),
        lambda: _score_src_contains_sq_stem(src_name, sq_connected),
        lambda: _score_lookup(src_name, lookup_source_names),
        lambda: _score_token_overlap(src_name, sq_connected),
    ]
    for scorer in scorers:
        result = scorer()
        if result is not None:
            return result
    return (ManifestConfidence.UNMAPPED, "No matching SQ, direct connector, or Lookup reference found")


def _mapping_for_source(src_name: str, connectors: list[dict], mapping_names: list[str]) -> str:
    for c in connectors:
        if c.get("from_instance") == src_name:
            return mapping_names[0] if mapping_names else "unknown"
    return mapping_names[0] if mapping_names else "unknown"


def _first_mapping_name(graph: dict) -> str:
    """Return the first mapping name in the graph, or 'unknown'."""
    names = [m["name"] for m in graph.get("mappings", [])]
    return names[0] if names else "unknown"


def _mapping_for_transform(tname: str, graph: dict) -> str:
    for m in graph.get("mappings", []):
        for t in m.get("transformations", []):
            if t["name"] == tname:
                return m["name"]
    return _first_mapping_name(graph)


# ─────────────────────────────────────────────────────────────────────────────
# Internal helpers — xlsx sheet builders
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(cell, text: str) -> None:
    cell.value = cell.value if text is None else text
    cell.font  = Font(name="Arial", bold=True, color=_WHITE, size=10)
    cell.fill  = PatternFill("solid", fgColor=_BLUE_H)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="AAAAAA")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _data_style(cell, fill_hex: Optional[str] = None, bold: bool = False,
                txt_color: str = "000000", wrap: bool = True) -> None:
    cell.font      = Font(name="Arial", size=9, bold=bold, color=txt_color)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)
    thin = Side(style="thin", color="DDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _conf_fill(conf: ManifestConfidence) -> str:
    return {
        ManifestConfidence.HIGH:     _GREEN,
        ManifestConfidence.MEDIUM:   _AMBER,
        ManifestConfidence.LOW:      _YELLOW,
        ManifestConfidence.UNMAPPED: "FFB3B3",
    }.get(conf, _WHITE)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_title(ws, report: ManifestReport) -> None:
    """Write the title and sub-title merged rows to the summary sheet."""
    ws.merge_cells("A1:G1")
    title = ws["A1"]
    title.value = "Informatica Mapping Manifest — Pre-Conversion Review"
    title.font  = Font(name="Arial", bold=True, size=13, color=_WHITE)
    title.fill  = PatternFill("solid", fgColor=_BLUE_H)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    review_flag = "YES ⚠️" if report.review_required else "NO ✅"
    sub.value = (
        f"Generated: {report.generated_at[:19].replace('T', ' ')} UTC"
        f"   |   Review Required: {review_flag}"
    )
    sub.font  = Font(name="Arial", italic=True, size=9, color="444444")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


def _write_summary_headers(ws) -> None:
    """Write the column header row (row 4) to the summary sheet."""
    headers = ["Mapping", "Sources", "Targets", "Transformations",
               "HIGH ✅", "MEDIUM ⚠", "LOW / UNMAPPED ❌"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[4].height = 20


def _build_summary_row(mname: str, report: ManifestReport) -> tuple[list, str]:
    """Compute the data values and fill colour for one summary row."""
    mapping_items = [x for x in report.items if x.mapping_name == mname]
    mc = _count_confidences(mapping_items)
    low_unmapped = mc[ManifestConfidence.LOW] + mc[ManifestConfidence.UNMAPPED]
    row_fill = _YELLOW if low_unmapped > 0 else _GREEN
    row_data = [
        mname,
        report.source_count,
        report.target_count,
        report.transformation_count,
        mc[ManifestConfidence.HIGH],
        mc[ManifestConfidence.MEDIUM],
        low_unmapped,
    ]
    return row_data, row_fill


def _write_summary_data_rows(ws, report: ManifestReport) -> None:
    """Write one data row per mapping into the summary sheet."""
    for i, mname in enumerate(report.mapping_names, start=5):
        row_data, row_fill = _build_summary_row(mname, report)
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            _data_style(cell, fill_hex=row_fill, bold=(col == 1))


def _build_summary_sheet(wb: Workbook, report: ManifestReport) -> None:
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False
    _write_summary_title(ws, report)
    _write_summary_headers(ws)
    _write_summary_data_rows(ws, report)
    _col_widths(ws, [40, 10, 10, 18, 10, 12, 18])


# ── Full Lineage sheet ────────────────────────────────────────────────────────

def _lineage_text_color(confidence: ManifestConfidence) -> str:
    """Return the text colour for a lineage row based on confidence."""
    if confidence in _REVIEW_CONFIDENCES:
        return _RED_TXT
    return "000000"


def _lineage_row_data(item: ManifestItem) -> list:
    """Build the cell values for one lineage row."""
    return [
        item.mapping_name,
        item.item_type.value,
        item.location,
        item.description,
        item.tool_determination,
        item.confidence.value,
        item.notes or "",
    ]


def _write_lineage_rows(ws, report: ManifestReport) -> None:
    """Write all data rows to the Full Lineage sheet."""
    lineage_items = [x for x in report.items if x.item_type in _LINEAGE_ITEM_TYPES]
    for row_idx, item in enumerate(lineage_items, start=2):
        fill = _conf_fill(item.confidence)
        txt  = _lineage_text_color(item.confidence)
        for col, val in enumerate(_lineage_row_data(item), start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, fill_hex=fill, txt_color=txt)


def _build_lineage_sheet(wb: Workbook, report: ManifestReport) -> None:
    ws = wb.create_sheet("Full Lineage")
    ws.sheet_view.showGridLines = False

    headers = ["Mapping", "Item Type", "Location", "Description",
               "Tool Determination", "Confidence", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    _write_lineage_rows(ws, report)
    _col_widths(ws, [30, 18, 28, 50, 45, 12, 35])


# ── Review Required sheet ─────────────────────────────────────────────────────

def _write_review_banner(ws) -> None:
    """Write the action banner at row 1 of the Review Required sheet."""
    ws.merge_cells("A1:H1")
    banner = ws["A1"]
    banner.value = (
        "⚠️  REVIEWER ACTION REQUIRED — Fill in the 'Reviewer Override' column "
        "for each row below, then save and re-upload."
    )
    banner.font  = Font(name="Arial", bold=True, size=10, color="9C0006")
    banner.fill  = PatternFill("solid", fgColor="FFE0E0")
    banner.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22


def _write_review_headers(ws) -> None:
    """Write the column header row (row 2) and highlight the Override column."""
    headers = ["Mapping", "Item Type", "Location", "Description",
               "Tool Determination", "Confidence", "Reviewer Override", "Notes"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=h)
        _header_style(cell, h)
    ws.row_dimensions[2].height = 20

    override_header = ws.cell(row=2, column=7)
    override_header.fill = PatternFill("solid", fgColor="C00000")
    override_header.font = Font(name="Arial", bold=True, color=_WHITE, size=10)


def _write_no_review_placeholder(ws) -> None:
    """Write the 'no review needed' message at row 3."""
    ws.merge_cells("A3:H3")
    cell = ws.cell(
        row=3, column=1,
        value="✅  No review required — all items resolved at HIGH or MEDIUM confidence.",
    )
    cell.font      = Font(name="Arial", italic=True, size=9, color="006100")
    cell.fill      = PatternFill("solid", fgColor=_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _review_item_row_data(item: ManifestItem) -> list:
    """Build the cell values for one review row."""
    return [
        item.mapping_name,
        item.item_type.value,
        item.location,
        item.description,
        item.tool_determination,
        item.confidence.value,
        item.reviewer_override or "",
        item.notes or "",
    ]


def _write_review_data_rows(ws, review_items: list) -> None:
    """Write item rows to the Review Required sheet."""
    for row_idx, item in enumerate(review_items, start=3):
        fill     = _YELLOW if item.confidence == ManifestConfidence.LOW else "FFB3B3"
        row_data = _review_item_row_data(item)
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            _data_style(cell, fill_hex=fill, bold=(col == 7))


def _build_review_sheet(wb: Workbook, report: ManifestReport) -> None:
    """
    The action sheet — only LOW/UNMAPPED items from ALL item types.
    Reviewer fills in the 'Reviewer Override' column and saves.
    """
    ws = wb.create_sheet("Review Required")
    ws.sheet_view.showGridLines = False
    _write_review_banner(ws)
    _write_review_headers(ws)
    ws.freeze_panes = "A3"

    review_items = [x for x in report.items if x.confidence in _REVIEW_CONFIDENCES]
    if not review_items:
        _write_no_review_placeholder(ws)
    else:
        _write_review_data_rows(ws, review_items)

    _col_widths(ws, [30, 18, 28, 50, 45, 12, 35, 35])
