# Copyright (c) 2026 ad25343 — https://github.com/ad25343/InformaticaConversion
# Licensed under CC BY-NC 4.0. Commercial use requires written permission.
"""
Source-to-Target Mapping Agent
Runs after Step 1 (Parse). Fully deterministic — no Claude needed.

Traces field lineage through the connector graph to produce:
  - A structured JSON S2T mapping (stored in job state)
  - An Excel workbook (saved to disk, downloadable from UI)

S2T record columns (one row per target field):
  Mapping | Source Table | Source Field | Source Type |
  Transformation Chain | Logic | Logic Type |
  Target Table | Target Field | Target Type |
  Status | Notes
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional
import re

from ..models.schemas import ParseReport
from .base import BaseAgent


# ── Output directory ─────────────────────────────────────────────────────
APP_DIR      = Path(__file__).parent.parent.parent
S2T_DIR      = APP_DIR / "logs" / "s2t"
S2T_DIR.mkdir(parents=True, exist_ok=True)

# ── Status values ────────────────────────────────────────────────────────
STATUS_DIRECT    = "Direct"        # source field → target with no transformation
STATUS_DERIVED   = "Derived"       # expression / derivation applied
STATUS_FILTERED  = "Filtered"      # passes through a Filter or Router
STATUS_LOOKUP    = "Lookup"        # enriched from a Lookup transformation
STATUS_AGGREGATE = "Aggregated"    # through an Aggregator
STATUS_UNMAPPED_SRC = "Unmapped Source"  # source field has no downstream target
STATUS_UNMAPPED_TGT = "Unmapped Target"  # target field has no upstream source


class S2TAgent(BaseAgent):

    def build_s2t(
        self,
        parse_report: ParseReport,
        graph: dict,
        job_id: str,
    ) -> dict:
        return _build_s2t_impl(parse_report, graph, job_id)


# ─────────────────────────────────────────────────────────────────────────────
# Connector index builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_backward_index(connectors: list[dict]) -> dict[tuple, tuple]:
    """Build (to_instance, to_field) → (from_instance, from_field) index."""
    return {
        (conn["to_instance"], conn["to_field"]): (conn["from_instance"], conn["from_field"])
        for conn in connectors
    }


def _build_forward_index(connectors: list[dict]) -> dict[tuple, list]:
    """Build (from_instance, from_field) → [(to_instance, to_field)] index."""
    forward: dict[tuple, list] = {}
    for conn in connectors:
        key = (conn["from_instance"], conn["from_field"])
        forward.setdefault(key, []).append((conn["to_instance"], conn["to_field"]))
    return forward


def _make_trans_getter(trans_by_name: dict, instance_map: dict):
    """Return a closure that resolves instance names to transformation dicts."""
    def get_trans(inst_name: str) -> Optional[dict]:
        if inst_name in trans_by_name:
            return trans_by_name[inst_name]
        mapped = instance_map.get(inst_name)
        if mapped and mapped in trans_by_name:
            return trans_by_name[mapped]
        return None
    return get_trans


# ─────────────────────────────────────────────────────────────────────────────
# Unmapped-source detection
# ─────────────────────────────────────────────────────────────────────────────

def _get_instance_sets(connectors: list[dict]) -> tuple[set, set]:
    """Return (to_instances, root_instances) from connector list."""
    to_instances   = {conn["to_instance"] for conn in connectors}
    root_instances = {conn["from_instance"] for conn in connectors} - to_instances
    return to_instances, root_instances


def _root_fields_used(connectors: list[dict], root_instances: set) -> set[tuple]:
    """Return (instance, field) pairs used by root (source) instances."""
    return {
        (conn["from_instance"], conn["from_field"])
        for conn in connectors
        if conn["from_instance"] in root_instances
    }


def _root_fields_connected(forward: dict, root_instances: set) -> set[tuple]:
    """Return forward-index keys belonging to root instances."""
    return {key for key in forward if key[0] in root_instances}


def _find_unmapped_source_fields(
    connectors: list[dict],
    graph: dict,
    mapping_name: str,
) -> list[dict]:
    """Return source fields that appear in root instances but never reach a target."""
    _to_instances, root_instances = _get_instance_sets(connectors)

    forward = _build_forward_index(connectors)
    used      = _root_fields_used(connectors, root_instances)
    connected = _root_fields_connected(forward, root_instances)

    unmapped: list[dict] = []
    for (src_inst, src_field) in used - connected:
        src_type = _find_source_field_type(src_field, graph)
        unmapped.append({
            "mapping_name": mapping_name,
            "source_table": src_inst,
            "source_field": src_field,
            "source_type":  src_type,
            "note": "Field present in source but not connected downstream",
        })
    return unmapped


def _find_source_field_type(src_field: str, graph: dict) -> str:
    """Find the datatype of a source field by scanning all graph sources."""
    for src in graph.get("sources", []):
        for sf in src.get("fields", []):
            if sf["name"] == src_field:
                return sf.get("datatype", "")
    return ""


# ─────────────────────────────────────────────────────────────────────────────
# Target-field record builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_target_record(
    mapping_name: str,
    tgt_name: str,
    tgt_field: str,
    tgt_type: str,
    result: dict,
    source_lookup: dict,
) -> dict:
    """Build a single S2T mapping record for a target field."""
    chain_ordered = list(reversed(result["chain"]))
    return {
        "mapping_name":        mapping_name,
        "source_table":        result["source_table"],
        "source_field":        result["source_field"],
        "source_type":         _get_source_field_type(
                                  result["source_table"],
                                  result["source_field"],
                                  source_lookup),
        "transformation_chain": chain_ordered,
        "transformation_chain_str": " → ".join(chain_ordered) if chain_ordered else "—",
        "logic":               result["logic"],
        "logic_type":          result["logic_type"],
        "target_table":        tgt_name,
        "target_field":        tgt_field,
        "target_type":         tgt_type,
        "status":              result["status"],
        "notes":               result["notes"],
    }


def _build_lkp_resolution(graph: dict) -> dict[str, str]:
    """
    Build a mapping of Lookup transformation instance names → actual lookup table names.

    Lookup transformations read from an external table (the lookup table) but have no
    CONNECTOR from that table — so the backward trace dead-ends at the LKP instance.
    This map lets us resolve LKP instances to their real source table when dead-ending.

    The lookup table name comes from the 'Lookup Table' or 'Lookup table name' attribute
    in the transformation definition.
    """
    lkp_map: dict[str, str] = {}
    for mapping in graph.get("mappings", []):
        for t in mapping.get("transformations", []):
            if "lookup" not in t.get("type", "").lower():
                continue
            inst = t["name"]
            attribs = t.get("table_attribs", {})
            lkp_table = (attribs.get("Lookup Table") or
                         attribs.get("Lookup table name") or
                         attribs.get("Lookup Table Name") or "").strip()
            if lkp_table:
                lkp_map[inst] = lkp_table
    return lkp_map


def _build_sq_resolution(graph: dict, source_names: set) -> dict[str, str]:
    """
    Build a mapping of Source Qualifier instance names → actual source table names.

    In Informatica, Source Qualifiers (SQ_*) read from SOURCE definitions but are
    not connected to them via CONNECTOR elements — so the backward trace can reach
    a SQ and find no further connectors.  This map lets us resolve SQ instances to
    their real source table when the trace dead-ends at one.

    Strategy: strip common SQ prefixes (SQ_, SQI_) and check against source_names.
    """
    sq_map: dict[str, str] = {}
    sq_prefixes = ("SQ_", "SQI_", "SRC_SQ_")
    source_upper = {s.upper(): s for s in source_names}

    for mapping in graph.get("mappings", []):
        for t in mapping.get("transformations", []):
            if "Source Qualifier" not in t.get("type", ""):
                continue
            inst = t["name"]
            for pfx in sq_prefixes:
                if not inst.upper().startswith(pfx):
                    continue
                candidate = inst[len(pfx):]
                cu = candidate.upper()

                # 1) Exact match (e.g. SQ_ORDERS → ORDERS)
                if cu in source_upper:
                    sq_map[inst] = source_upper[cu]
                    break

                # 2) Suffix match: source may have an extra schema prefix
                #    e.g. SQ_LOAN_APPLICATIONS → LOAN_APPLICATIONS
                #         matches STG_LOAN_APPLICATIONS (ends with _LOAN_APPLICATIONS)
                for sname_upper, sname in source_upper.items():
                    if sname_upper.endswith("_" + cu):
                        sq_map[inst] = sname
                        break

                if inst in sq_map:
                    break
    return sq_map


def _build_target_instance_map(backward: dict, target_names: set) -> dict[str, str]:
    """
    Some mappings use a 'TGT_' prefixed instance name for target definitions
    (e.g. instance TGT_FACT_ORDERS connects to target definition FACT_ORDERS).
    Build a map: target_definition_name → instance_name_in_connectors so that
    the trace starts at the right key in the backward index.
    """
    # Collect all instance names that appear as to_instance in the backward index
    to_instances = {k[0] for k in backward}
    result: dict[str, str] = {}
    for tname in target_names:
        if tname in to_instances:
            result[tname] = tname          # instance name matches definition name
        elif f"TGT_{tname}" in to_instances:
            result[tname] = f"TGT_{tname}" # TGT_ prefix convention
        else:
            # Fallback: try case-insensitive prefix match
            upper = tname.upper()
            for inst in to_instances:
                if inst.upper().endswith(upper) or upper.endswith(inst.upper().lstrip("TGT_")):
                    result[tname] = inst
                    break
    return result


def _process_mapping_targets(
    mapping: dict,
    graph: dict,
    source_lookup: dict,
    backward: dict,
    source_names: set,
    trans_by_name: dict,
    get_trans,
    sq_resolution: dict,
    lkp_resolution: dict | None = None,
) -> tuple[list[dict], list[dict]]:
    """
    Trace each target field in a mapping to its source.
    Returns (records, unmapped_targets).
    """
    mapping_name = mapping["name"]
    records: list[dict] = []
    unmapped_targets: list[dict] = []

    target_names = {tgt["name"] for tgt in graph.get("targets", [])}
    tgt_instance_map = _build_target_instance_map(backward, target_names)

    for tgt in graph.get("targets", []):
        tgt_name = tgt["name"]
        # Resolve instance name (may have TGT_ prefix in connectors)
        start_instance = tgt_instance_map.get(tgt_name, tgt_name)
        for tgt_field_def in tgt.get("fields", []):
            tgt_field = tgt_field_def["name"]
            tgt_type  = tgt_field_def.get("datatype", "")

            result = _trace_to_source(
                start_instance, tgt_field,
                backward, source_names, trans_by_name, get_trans,
                sq_resolution=sq_resolution,
                lkp_resolution=lkp_resolution,
            )

            if result["source_table"] is None:
                unmapped_targets.append({
                    "mapping_name": mapping_name,
                    "target_table": tgt_name,
                    "target_field": tgt_field,
                    "target_type":  tgt_type,
                    "note": "No upstream connector found",
                })
                continue

            records.append(_build_target_record(
                mapping_name, tgt_name, tgt_field, tgt_type, result, source_lookup
            ))

    return records, unmapped_targets


def _count_records_by_status(records: list[dict], status: str) -> int:
    """Count records matching a single status value."""
    return sum(1 for r in records if r["status"] == status)


def _build_summary(records: list[dict], unmapped_targets: list[dict], unmapped_sources: list[dict]) -> dict:
    """Build the summary statistics dict."""
    return {
        "total_target_fields":    len(records) + len(unmapped_targets),
        "mapped_fields":          len(records),
        "unmapped_target_fields": len(unmapped_targets),
        "unmapped_source_fields": len(unmapped_sources),
        "direct_mappings":        _count_records_by_status(records, STATUS_DIRECT),
        "derived_mappings":       _count_records_by_status(records, STATUS_DERIVED),
        "lookup_enriched":        _count_records_by_status(records, STATUS_LOOKUP),
        "filtered_fields":        _count_records_by_status(records, STATUS_FILTERED),
    }


def _build_s2t_impl(
    parse_report: ParseReport,
    graph: dict,
    job_id: str,
) -> dict:
    """
    Build the full S2T mapping for a job.

    Returns a dict:
      {
        "records":        [S2TRecord, ...],       # one per target field
        "unmapped_sources": [...],                # source fields with no target
        "unmapped_targets": [...],                # target fields with no source
        "summary": {...},
        "excel_path": "logs/s2t/<name>.xlsx",     # relative path for download
      }
    """
    records: list[dict] = []
    unmapped_sources: list[dict] = []
    unmapped_targets: list[dict] = []

    source_lookup  = {s["name"]: s for s in graph.get("sources", [])}
    source_names   = set(source_lookup.keys())

    # Build resolution maps (once, across all mappings)
    sq_resolution  = _build_sq_resolution(graph, source_names)
    lkp_resolution = _build_lkp_resolution(graph)

    for mapping in graph.get("mappings", []):
        connectors   = mapping.get("connectors", [])
        instance_map = mapping.get("instance_map", {})

        trans_by_name: dict[str, dict] = {
            t["name"]: t for t in mapping.get("transformations", [])
        }
        get_trans = _make_trans_getter(trans_by_name, instance_map)
        backward  = _build_backward_index(connectors)

        m_records, m_unmapped_tgt = _process_mapping_targets(
            mapping, graph, source_lookup, backward, source_names,
            trans_by_name, get_trans, sq_resolution, lkp_resolution,
        )
        records.extend(m_records)
        unmapped_targets.extend(m_unmapped_tgt)
        unmapped_sources.extend(
            _find_unmapped_source_fields(connectors, graph, mapping["name"])
        )

    # ── Build Excel ───────────────────────────────────────────────────────
    mapping_stem   = parse_report.mapping_names[0] if parse_report.mapping_names else job_id[:8]
    excel_filename = f"{_safe(mapping_stem)}_s2t_{job_id[:8]}.xlsx"
    excel_path     = S2T_DIR / excel_filename
    _write_excel(records, unmapped_sources, unmapped_targets, excel_path, mapping_stem)

    return {
        "records":            records,
        "unmapped_sources":   unmapped_sources,
        "unmapped_targets":   unmapped_targets,
        "summary":            _build_summary(records, unmapped_targets, unmapped_sources),
        "excel_filename":     excel_filename,
        "excel_path":         str(excel_path),
    }


# Backward-compat shim — keeps orchestrator.py call sites unchanged
def build_s2t(
    parse_report: ParseReport,
    graph: dict,
    job_id: str,
) -> dict:
    return S2TAgent().build_s2t(parse_report, graph, job_id)


# ─────────────────────────────────────────────────────────────────────────────
# Lineage tracer
# ─────────────────────────────────────────────────────────────────────────────

def _trace_result(
    source_table, source_field, chain, logic, logic_type, status, notes
) -> dict:
    """Build a trace result dict from components."""
    return {
        "source_table": source_table,
        "source_field": source_field,
        "chain": chain,
        "logic": "; ".join(logic) if isinstance(logic, list) else logic,
        "logic_type": logic_type,
        "status": status,
        "notes": "; ".join(notes) if isinstance(notes, list) else notes,
    }


def _no_upstream_result(chain: list, status: str) -> dict:
    """Return the unmapped-target result when no upstream connector exists."""
    return _trace_result(None, None, chain, "", status, STATUS_UNMAPPED_TGT, "No upstream connector found")


def _found_source_result(from_inst: str, from_field: str, chain: list, logic: list, status: str, notes: list) -> dict:
    """Return a successful trace result when a named source table was reached."""
    return _trace_result(from_inst, from_field, chain, logic, status, status, notes)


def _find_expr_for_port(trans: dict, port: str) -> str:
    """Return the expression text for a named port, or empty string if not found.

    Handles Router group-qualified names (e.g., FLAGGED_FRAUD_SCORE → FRAUD_SCORE)
    by also trying the port name with the group prefix stripped.
    """
    for expr in trans.get("expressions", []):
        if expr["port"] == port:
            return expr["expression"]
    # Router group-qualified names: strip prefix and retry
    # e.g., FLAGGED_FRAUD_SCORE → try FRAUD_SCORE, CLEARED_TRANSACTION_ID → TRANSACTION_ID
    if "_" in port:
        # Try progressively stripping prefixes (handle multi-word group names)
        parts = port.split("_")
        for i in range(1, len(parts)):
            stripped = "_".join(parts[i:])
            for expr in trans.get("expressions", []):
                if expr["port"] == stripped:
                    return expr["expression"]
    return ""


def _follow_internal_expr_chain(
    trans: dict,
    start_port: str,
    current_inst: str,
    backward: dict,
    max_depth: int = 5,
) -> str:
    """
    Walk the expression chain *within* a single transformation to find an input
    port that has a backward connector.

    Handles multi-level internal derivations such as:
        RISK_BAND   → IIF(FRAUD_SCORE >= 80, ...)
        FRAUD_SCORE → ROUND(INDICATOR_SCORE * PATTERN_WEIGHT, 2)
        INDICATOR_SCORE has backward[(EXP, INDICATOR_SCORE)] = (JNR, INDICATOR_SCORE) ← found

    Returns the first discoverable input port name, or "" if none found.
    """
    port_names = {p["name"] for p in trans.get("ports", [])}
    visited: set[str] = {start_port}
    queue: list[str] = [start_port]

    for _ in range(max_depth):
        if not queue:
            break
        next_queue: list[str] = []
        for port in queue:
            expr = _find_expr_for_port(trans, port)
            if not expr or expr == port:
                continue
            tokens = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", expr)
            for token in tokens:
                if token in visited:
                    continue
                if token in port_names:
                    if (current_inst, token) in backward:
                        return token          # found an input with a backward connector
                    visited.add(token)
                    next_queue.append(token)  # internal derived port — explore further
        queue = next_queue
    return ""


def _find_followable_token(
    expr_text: str,
    current_field: str,
    current_inst: str,
    port_names: set,
    backward: dict,
    notes: list,
    trans: dict | None = None,
) -> str:
    """
    Scan expression tokens for one that has a backward connector.

    When a token is an internal derived port (in port_names but not in backward),
    the expression chain within the transformation is followed recursively to find
    an upstream input port that does have a backward connector.

    Returns the resolved port name on success, or "" if none found.
    Appends to notes when a redirect is found.
    """
    tokens = re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", expr_text)
    for token in tokens:
        if token == current_field:
            continue
        if token not in port_names:
            continue
        if (current_inst, token) in backward:
            notes.append(
                f"'{current_field}' derived via expression "
                f"({_truncate(expr_text, 80)}); tracing through '{token}'"
            )
            return token
        # token is an internal derived port — follow its expression chain
        if trans is not None:
            resolved = _follow_internal_expr_chain(trans, token, current_inst, backward)
            if resolved:
                notes.append(
                    f"'{current_field}' derived via internal chain "
                    f"({_truncate(expr_text, 80)}) → '{token}' → '{resolved}'"
                )
                return resolved
    return ""


def _try_follow_expression(
    current_inst: str,
    current_field: str,
    backward: dict,
    chain: list,
    logic: list,
    notes: list,
    status: str,
    get_trans,
) -> Optional[str]:
    """
    If current node is a derived output port in a transformation with an expression,
    try to find an upstream input port to follow instead.

    Returns the new field name to follow (non-empty string), "" if root (expression
    found but no followable token), or None if the current node has no transformation.
    """
    trans = get_trans(current_inst)
    if trans is None:
        return None  # true root

    expr_text = _find_expr_for_port(trans, current_field)
    if not expr_text or expr_text == current_field:
        return None  # no expression, treat as root

    port_names = {p["name"] for p in trans.get("ports", [])}
    return _find_followable_token(expr_text, current_field, current_inst, port_names, backward, notes, trans=trans)


_STATUS_FOR_TTYPE: dict[str, str] = {
    "Aggregator":        STATUS_AGGREGATE,
    "Filter":            STATUS_FILTERED,
    "Router":            STATUS_FILTERED,
    "Lookup Procedure":  STATUS_LOOKUP,
}


def _is_lookup_type(ttype: str) -> bool:
    return "Lookup" in ttype


def _promote_status(current: str, candidate: str) -> str:
    """Return candidate only if current is still STATUS_DIRECT, otherwise keep current."""
    return candidate if current == STATUS_DIRECT else current


def _apply_transformation_status(
    ttype: str,
    status: str,
    from_inst: str,
    notes: list,
) -> str:
    """Update status based on transformation type; append notes for Router/Joiner."""
    if _is_lookup_type(ttype):
        return _promote_status(status, STATUS_LOOKUP)
    if ttype == "Router":
        notes.append(f"Routes through {from_inst}")
        return _promote_status(status, STATUS_FILTERED)
    if ttype == "Joiner":
        notes.append(f"Joined at {from_inst}")
        return status
    new_status = _STATUS_FOR_TTYPE.get(ttype)
    if new_status:
        return _promote_status(status, new_status)
    return status


def _collect_port_logic(trans: dict, from_inst: str, from_field: str, logic: list) -> str:
    """
    If the transformation has a non-trivial expression for from_field,
    append it to logic and return STATUS_DERIVED; else return STATUS_DIRECT.
    """
    expr_text = _find_expr_for_port(trans, from_field)
    if expr_text and expr_text != from_field:
        logic.append(f"{from_inst}.{from_field}: {_truncate(expr_text, 120)}")
        return STATUS_DERIVED
    return STATUS_DIRECT


def _process_intermediate_node(
    from_inst: str,
    from_field: str,
    trans_by_name: dict,
    get_trans,
    chain: list,
    logic: list,
    notes: list,
    status: str,
) -> str:
    """
    Process an intermediate transformation node encountered during backward trace.
    Mutates chain, logic, notes in place and returns the updated status.
    """
    chain.append(from_inst)

    trans = get_trans(from_inst)
    if not trans:
        return status

    ttype = trans.get("type", "")
    new_status = _collect_port_logic(trans, from_inst, from_field, logic)
    if new_status == STATUS_DERIVED:
        status = STATUS_DERIVED

    return _apply_transformation_status(ttype, status, from_inst, notes)


def _handle_dead_end(
    current_inst: str,
    current_field: str,
    hops: int,
    backward: dict,
    chain: list,
    logic: list,
    notes: list,
    status: str,
    get_trans,
) -> tuple[Optional[dict], Optional[str]]:
    """
    Handle the case where (current_inst, current_field) has no backward connector.

    Returns (result_dict, None) if we've reached a terminal node,
    or (None, new_field) if we should redirect and continue tracing.
    """
    if hops == 0:
        # Even at hops==0, check for OUTPUT-only expressions (e.g., SYSDATE, $$param)
        trans = get_trans(current_inst)
        if trans:
            expr_text = _find_expr_for_port(trans, current_field)
            if expr_text and expr_text != current_field:
                logic.append(f"{current_inst}.{current_field}: {_truncate(expr_text, 120)}")
                return _trace_result(
                    None, None, chain, "; ".join(logic), STATUS_DERIVED,
                    STATUS_DERIVED, f"Computed field — {_truncate(expr_text, 60)}"
                ), None
        return _no_upstream_result(chain, status), None

    # Try to find and follow an expression token upstream
    new_field = _try_follow_expression(
        current_inst, current_field, backward, chain, logic, notes, status, get_trans
    )
    if new_field:
        return None, new_field  # redirect to follow this expression token

    # No followable token — but still capture the expression as derivation logic
    trans = get_trans(current_inst)
    if trans:
        expr_text = _find_expr_for_port(trans, current_field)
        if expr_text and expr_text != current_field:
            logic.append(f"{current_inst}.{current_field}: {_truncate(expr_text, 120)}")
            status = STATUS_DERIVED

    return _found_source_result(current_inst, current_field, chain, logic, status, notes), None


def _trace_to_source(
    start_instance: str,
    start_field: str,
    backward: dict,
    source_names: set,
    trans_by_name: dict,
    get_trans,
    max_depth: int = 20,
    sq_resolution: dict | None = None,
    lkp_resolution: dict | None = None,
) -> dict:
    """
    Walk the backward connector graph from a target field to its ultimate source.
    Returns a dict with: source_table, source_field, chain, logic, logic_type, status, notes.

    sq_resolution:  maps Source Qualifier instance names → actual source table names.
    lkp_resolution: maps Lookup transformation names → their lookup table names.
    Both are applied when the trace dead-ends at one of these transformation types.
    """
    if sq_resolution is None:
        sq_resolution = {}
    if lkp_resolution is None:
        lkp_resolution = {}

    chain:  list[str] = []
    logic:  list[str] = []
    notes:  list[str] = []
    status = STATUS_DIRECT

    current_inst  = start_instance
    current_field = start_field
    hops = 0

    # Informatica Joiner port-name conventions (multiple styles exist in the wild):
    #
    #   Style A — M_/D_ prefix on INPUT:
    #     SQ → JNR input port  M_INDICATOR_ID  (master)
    #     JNR output port  INDICATOR_ID        (no prefix)
    #     → at (JNR, INDICATOR_ID) try M_INDICATOR_ID in backward
    #
    #   Style B — OUT_ prefix on OUTPUT + optional _M/_D suffix on ambiguous inputs:
    #     SQ → JNR input port  CUSTOMER_ID_M   (master, only on clashing names)
    #     SQ → JNR input port  APPLICATION_ID  (no suffix — field exists in one side only)
    #     JNR output port  OUT_CUSTOMER_ID / OUT_APPLICATION_ID
    #     → at (JNR, OUT_CUSTOMER_ID):
    #         1. strip OUT_ → CUSTOMER_ID
    #         2. check (JNR, CUSTOMER_ID) in backward (works for non-ambiguous fields)
    #         3. try _M / _D suffixes (works for ambiguous ones)
    _JOINER_PREFIXES = ("M_", "MASTER_", "D_", "DETAIL_")
    _JOINER_SUFFIXES = ("_M", "_D", "_MASTER", "_DETAIL")

    for _ in range(max_depth):
        key = (current_inst, current_field)
        if key not in backward:
            resolved = False
            trans = get_trans(current_inst)
            ttype = trans.get("type", "") if trans else ""

            # Router group-qualified ports: strip prefix to find the INPUT port
            # e.g., (RTR_X, FLAGGED_FRAUD_SCORE) → try (RTR_X, FRAUD_SCORE)
            if ttype == "Router" and "_" in current_field:
                parts = current_field.split("_")
                for i in range(1, len(parts)):
                    stripped = "_".join(parts[i:])
                    alt_key = (current_inst, stripped)
                    if alt_key in backward:
                        if trans:
                            _collect_port_logic(trans, current_inst, current_field, logic)
                        current_field = stripped
                        resolved = True
                        break

            # Joiner port disambiguation — handles two naming conventions:
            elif ttype == "Joiner":
                # Style B: strip leading OUT_ from output port, then resolve
                base_field = (current_field[4:] if current_field.upper().startswith("OUT_")
                              else current_field)

                # 1) Base name directly (non-ambiguous fields, no disambiguation suffix)
                if base_field != current_field:
                    alt_key = (current_inst, base_field)
                    if alt_key in backward:
                        current_field = base_field
                        resolved = True

                # 2) Style A: M_/D_/MASTER_/DETAIL_ prefix on input port
                if not resolved:
                    for pfx in _JOINER_PREFIXES:
                        alt_key = (current_inst, pfx + base_field)
                        if alt_key in backward:
                            current_field = pfx + base_field
                            resolved = True
                            break

                # 3) Style B: _M/_D/_MASTER/_DETAIL suffix on input port
                if not resolved:
                    for sfx in _JOINER_SUFFIXES:
                        alt_key = (current_inst, base_field + sfx)
                        if alt_key in backward:
                            current_field = base_field + sfx
                            resolved = True
                            break

            if resolved:
                continue

            result, new_field = _handle_dead_end(
                current_inst, current_field, hops, backward, chain, logic, notes, status, get_trans
            )
            if result is not None:
                # If we dead-ended on a Source Qualifier, resolve to real source name
                if (result.get("source_table") and
                        result["source_table"] in sq_resolution):
                    result = dict(result)
                    result["source_table"] = sq_resolution[result["source_table"]]
                # Lookup dead-end resolution
                elif (result.get("source_table") and
                        result["source_table"] in lkp_resolution):
                    result = dict(result)
                    result["source_table"] = lkp_resolution[result["source_table"]]
                return result
            current_field = new_field
            continue

        from_inst, from_field = backward[key]
        hops += 1

        if from_inst in source_names:
            return _found_source_result(from_inst, from_field, chain, logic, status, notes)

        # Source Qualifier resolution: SQ instances are roots (no further backward
        # connectors exist between SOURCE and SQ in Informatica XML).
        if from_inst in sq_resolution:
            return _found_source_result(sq_resolution[from_inst], from_field, chain, logic, status, notes)

        # Lookup resolution: Lookup output fields have no backward connector to
        # their source table — resolve via the Lookup Table attribute.
        if from_inst in lkp_resolution:
            return _found_source_result(lkp_resolution[from_inst], from_field, chain, logic, status, notes)

        status = _process_intermediate_node(
            from_inst, from_field, trans_by_name, get_trans, chain, logic, notes, status
        )
        current_inst  = from_inst
        current_field = from_field

    return _trace_result(
        None, None, chain,
        "; ".join(logic), status,
        "Trace Too Deep",
        f"Could not resolve — chain exceeds {max_depth} hops",
    )


def _get_source_field_type(src_table: str, src_field: str, source_lookup: dict) -> str:
    src = source_lookup.get(src_table, {})
    for f in src.get("fields", []):
        if f["name"] == src_field:
            return f.get("datatype", "")
    return ""


def _truncate(s: str, n: int) -> str:
    return s if len(s) <= n else s[:n] + "…"

def _safe(s: str) -> str:
    return re.sub(r"[^\w\-]", "_", s)[:50]


# ─────────────────────────────────────────────────────────────────────────────
# Excel writer
# ─────────────────────────────────────────────────────────────────────────────

# Colour palette constants — used by all sheet helpers
_C_HEADER_BG  = "1E293B"
_C_HEADER_FG  = "F8FAFC"
_C_SRC_BG     = "EFF6FF"
_C_TGT_BG     = "F0FDF4"
_C_DERIVED_BG = "FFFBEB"
_C_LOOKUP_BG  = "FAF5FF"
_C_FILTER_BG  = "FFF7ED"
_C_ERROR_BG   = "FEF2F2"
_C_DIRECT_BG  = "F0FDF4"
_C_ALT_BG     = "F8FAFC"
_C_ACCENT     = "6366F1"


def _make_excel_styles():
    """Import openpyxl style objects and return a tuple of commonly used ones."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_style() -> dict:
        return {
            "font":      Font(bold=True, color=_C_HEADER_FG, name="Calibri", size=10),
            "fill":      PatternFill("solid", fgColor=_C_HEADER_BG),
            "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
            "border":    border,
        }

    def apply(cell, **kwargs):
        from copy import copy
        for attr, val in kwargs.items():
            setattr(cell, attr, copy(val))

    def apply_header(cell, value):
        cell.value = value
        apply(cell, **header_style())

    status_fill = {
        STATUS_DIRECT:    PatternFill("solid", fgColor=_C_DIRECT_BG),
        STATUS_DERIVED:   PatternFill("solid", fgColor=_C_DERIVED_BG),
        STATUS_LOOKUP:    PatternFill("solid", fgColor=_C_LOOKUP_BG),
        STATUS_FILTERED:  PatternFill("solid", fgColor=_C_FILTER_BG),
        STATUS_AGGREGATE: PatternFill("solid", fgColor=_C_LOOKUP_BG),
        STATUS_UNMAPPED_TGT: PatternFill("solid", fgColor=_C_ERROR_BG),
        STATUS_UNMAPPED_SRC: PatternFill("solid", fgColor=_C_ERROR_BG),
    }

    return border, apply, apply_header, status_fill, Font, PatternFill, Alignment


_MAIN_COLS = [
    ("Mapping",              18),
    ("Source Table",         22),
    ("Source Field",         22),
    ("Source Type",          14),
    ("Transformation Chain", 35),
    ("Logic / Expression",   45),
    ("Logic Type",           14),
    ("Target Table",         22),
    ("Target Field",         22),
    ("Target Type",          14),
    ("Status",               14),
    ("Notes",                35),
]


def _write_sheet_title(ws, text: str, col_span: str, Font, PatternFill, Alignment, color: str = _C_ACCENT):
    """Write a merged title row to a worksheet."""
    ws.merge_cells(col_span)
    cell = ws[col_span.split(":")[0]]
    cell.value     = text
    cell.font      = Font(bold=True, size=13, color=color, name="Calibri")
    cell.fill      = PatternFill("solid", fgColor="F1F5F9")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28


def _write_col_headers(ws, cols: list, apply_header, get_column_letter, start_row: int = 2):
    """Write column headers and set column widths."""
    for col_idx, (col_name, col_width) in enumerate(cols, start=1):
        apply_header(ws.cell(row=start_row, column=col_idx), col_name)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[start_row].height = 22


_SRC_COLS = frozenset((2, 3, 4))
_TGT_COLS = frozenset((8, 9, 10))


def _src_fill_color(is_alt: bool) -> str:
    return "EBF4FF" if is_alt else _C_SRC_BG


def _tgt_fill_color(is_alt: bool) -> str:
    return "E8FAF0" if is_alt else _C_TGT_BG


def _get_cell_fill_for_col(col_idx: int, is_alt: bool, status_fill, r_status: str, PatternFill):
    """Return the appropriate fill for a data cell based on column index and row status."""
    default_fgColor = _C_ALT_BG if is_alt else "FFFFFF"
    default_fill = PatternFill("solid", fgColor=default_fgColor)
    if col_idx == 11:
        return status_fill.get(r_status, default_fill)
    if col_idx in _SRC_COLS:
        return PatternFill("solid", fgColor=_src_fill_color(is_alt))
    if col_idx in _TGT_COLS:
        return PatternFill("solid", fgColor=_tgt_fill_color(is_alt))
    return default_fill


def _write_data_row(ws, row_idx: int, values: list, r_status: str, border, status_fill, Font, PatternFill, Alignment):
    """Write a single data row to the main sheet."""
    from copy import copy
    is_alt = (row_idx % 2 == 0)
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border    = copy(border)
        cell.alignment = Alignment(vertical="top", wrap_text=(col_idx in (5, 6, 12)))
        cell_fill = _get_cell_fill_for_col(col_idx, is_alt, status_fill, r_status, PatternFill)
        cell.fill = copy(cell_fill)
        is_status_col = col_idx == 11
        cell.font = Font(name="Calibri", size=9, bold=is_status_col)
    ws.row_dimensions[row_idx].height = 16


def _write_legend(ws, legend_row: int, border, Font, PatternFill):
    """Write the status legend block below the data table."""
    from copy import copy
    ws.cell(row=legend_row, column=1).value = "Status Legend"
    ws.cell(legend_row, 1).font = Font(bold=True, name="Calibri", size=9)
    legend_items = [
        (STATUS_DIRECT,    _C_DIRECT_BG),
        (STATUS_DERIVED,   _C_DERIVED_BG),
        (STATUS_LOOKUP,    _C_LOOKUP_BG),
        (STATUS_FILTERED,  _C_FILTER_BG),
        (STATUS_AGGREGATE, _C_LOOKUP_BG),
        (STATUS_UNMAPPED_TGT, _C_ERROR_BG),
    ]
    for i, (status, fill_color) in enumerate(legend_items):
        cell = ws.cell(row=legend_row + 1 + i, column=1, value=status)
        cell.fill   = PatternFill("solid", fgColor=fill_color)
        cell.border = copy(border)
        cell.font   = Font(name="Calibri", size=9)


def _write_main_sheet(wb, records: list[dict], mapping_name: str, border, apply_header, status_fill, Font, PatternFill, Alignment):
    """Populate Sheet 1 — Field Mapping."""
    from openpyxl.utils import get_column_letter
    ws = wb.active
    ws.title = "Field Mapping"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A3"

    _write_sheet_title(ws, f"Source-to-Target Field Mapping — {mapping_name}", "A1:L1", Font, PatternFill, Alignment)
    _write_col_headers(ws, _MAIN_COLS, apply_header, get_column_letter)

    for row_idx, r in enumerate(records, start=3):
        values = [
            r["mapping_name"], r["source_table"], r["source_field"], r["source_type"],
            r["transformation_chain_str"], r["logic"], r["logic_type"],
            r["target_table"], r["target_field"], r["target_type"],
            r["status"], r["notes"],
        ]
        _write_data_row(ws, row_idx, values, r["status"], border, status_fill, Font, PatternFill, Alignment)

    if records:
        ws.auto_filter.ref = f"A2:L{len(records) + 2}"

    _write_legend(ws, len(records) + 5, border, Font, PatternFill)


_UNMAPPED_SRC_COLS = [
    ("Mapping", 22), ("Source Table", 24), ("Source Field", 24),
    ("Source Type", 14), ("Note", 50),
]
_UNMAPPED_TGT_COLS = [
    ("Mapping", 22), ("Target Table", 24), ("Target Field", 24),
    ("Target Type", 14), ("Note", 50),
]


def _write_unmapped_sheet(wb, sheet_name: str, title_text: str, col_defs: list, rows: list[dict], key_order: list, border, apply_header, Font, PatternFill, Alignment):
    """Write an unmapped sources/targets sheet."""
    from openpyxl.utils import get_column_letter
    from copy import copy
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    col_span = f"A1:{get_column_letter(len(col_defs))}1"
    ws.merge_cells(col_span)
    t = ws[col_span.split(":")[0]]
    t.value     = title_text
    t.font      = Font(bold=True, size=11, color="DC2626", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=_C_ERROR_BG)
    t.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    _write_col_headers(ws, col_defs, apply_header, get_column_letter)

    for row_idx, r in enumerate(rows, start=3):
        vals = [r[k] for k in key_order]
        for col_idx, val in enumerate(vals, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = copy(border)
            cell.fill      = PatternFill("solid", fgColor=_C_ERROR_BG if row_idx % 2 == 0 else "FFFFFF")
            cell.font      = Font(name="Calibri", size=9)
            cell.alignment = Alignment(vertical="top")


def _is_section_label(label: str) -> bool:
    return label.startswith("Total") or label.startswith("Mapped") or label.startswith("Unmapped")


def _count_by_status(records: list[dict], *statuses: str) -> int:
    """Count records whose status is one of the given statuses."""
    status_set = set(statuses)
    return sum(1 for r in records if r["status"] in status_set)


def _build_summary_rows(records: list[dict], unmapped_targets: list[dict], unmapped_sources: list[dict]) -> list[tuple]:
    """Build the (label, value) pairs for the summary sheet."""
    return [
        ("Total Target Fields",    len(records) + len(unmapped_targets)),
        ("Mapped Fields",          len(records)),
        ("  — Direct",             _count_by_status(records, STATUS_DIRECT)),
        ("  — Derived",            _count_by_status(records, STATUS_DERIVED)),
        ("  — Lookup Enriched",    _count_by_status(records, STATUS_LOOKUP)),
        ("  — Filtered / Routed",  _count_by_status(records, STATUS_FILTERED, STATUS_AGGREGATE)),
        ("Unmapped Target Fields", len(unmapped_targets)),
        ("Unmapped Source Fields", len(unmapped_sources)),
    ]


def _write_summary_row(ws4, i: int, label: str, val, border, Font, PatternFill, Alignment):
    """Write one row to the summary sheet."""
    from copy import copy
    lc = ws4.cell(row=i, column=1, value=label)
    vc = ws4.cell(row=i, column=2, value=val)
    lc.font  = Font(name="Calibri", size=10, bold=_is_section_label(label))
    vc.font  = Font(name="Calibri", size=10, bold=True)
    row_fill = PatternFill("solid", fgColor="F8FAFC" if i % 2 == 0 else "FFFFFF")
    lc.fill  = row_fill
    vc.fill  = copy(row_fill)
    lc.alignment = Alignment(vertical="center")
    vc.alignment = Alignment(horizontal="center", vertical="center")
    lc.border = copy(border)
    vc.border = copy(border)
    ws4.row_dimensions[i].height = 18
    if label.startswith("Unmapped") and val > 0:
        vc.font = Font(name="Calibri", size=10, bold=True, color="DC2626")


def _write_summary_sheet(wb, records: list[dict], unmapped_sources: list[dict], unmapped_targets: list[dict], mapping_name: str, border, Font, PatternFill, Alignment):
    """Write Sheet 4 — Summary (inserted at position 0)."""
    ws4 = wb.create_sheet("Summary", 0)
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 32
    ws4.column_dimensions["B"].width = 20

    ws4.merge_cells("A1:B1")
    s1 = ws4["A1"]
    s1.value     = f"S2T Summary — {mapping_name}"
    s1.font      = Font(bold=True, size=14, color=_C_ACCENT, name="Calibri")
    s1.fill      = PatternFill("solid", fgColor="EEF2FF")
    s1.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 32

    summary_rows = _build_summary_rows(records, unmapped_targets, unmapped_sources)
    for i, (label, val) in enumerate(summary_rows, start=3):
        _write_summary_row(ws4, i, label, val, border, Font, PatternFill, Alignment)


def _write_excel(
    records: list[dict],
    unmapped_sources: list[dict],
    unmapped_targets: list[dict],
    path: Path,
    mapping_name: str,
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment

    wb = openpyxl.Workbook()
    border, apply, apply_header, status_fill, Font, PatternFill, Alignment = _make_excel_styles()

    _write_main_sheet(wb, records, mapping_name, border, apply_header, status_fill, Font, PatternFill, Alignment)

    if unmapped_sources:
        _write_unmapped_sheet(
            wb, "Unmapped Sources",
            "Unmapped Source Fields — these fields exist in the source but are not used in any target mapping",
            _UNMAPPED_SRC_COLS,
            unmapped_sources,
            ["mapping_name", "source_table", "source_field", "source_type", "note"],
            border, apply_header, Font, PatternFill, Alignment,
        )

    if unmapped_targets:
        _write_unmapped_sheet(
            wb, "Unmapped Targets",
            "Unmapped Target Fields — these target columns have no mapped source",
            _UNMAPPED_TGT_COLS,
            unmapped_targets,
            ["mapping_name", "target_table", "target_field", "target_type", "note"],
            border, apply_header, Font, PatternFill, Alignment,
        )

    _write_summary_sheet(wb, records, unmapped_sources, unmapped_targets, mapping_name, border, Font, PatternFill, Alignment)

    wb.save(path)


# ─────────────────────────────────────────────────────────────────────────────
# Lookup helper — find S2T output path for a job
# ─────────────────────────────────────────────────────────────────────────────

def s2t_excel_path(job_id: str) -> Optional[Path]:
    """Find the S2T Excel file for a given job_id (by short ID suffix)."""
    for f in S2T_DIR.glob(f"*_{job_id[:8]}.xlsx"):
        return f
    return None
